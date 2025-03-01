import sqlite3
from telebot import types
import logging
import os
from PIL import Image, ImageDraw, ImageFont
import io
import pandas as pd
import openpyxl
from fpdf import FPDF
import random
import telebot
import math

logger = logging.getLogger(__name__)

# Определяем путь к базе данных один раз
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'db', 'characters.db')

# Обновленная карта характеристик с тематическими эмодзи
CHARACTERISTICS_MAP = {
    'hp': ('❤️ ОЗ', 'ОЗ'),
    'hp_regen': ('🩸 Реген ОЗ', 'реген_ОЗ'),
    'mana': ('💙 Мана', 'мана/энергия'),
    'mana_regen': ('💧 Реген маны', 'реген_маны/энергии'),
    'phys_attack': ('⚔️ Физ.атака', 'физ_атака'),
    'phys_def': ('🛡️ Физ.защита', 'физ_защита'),
    'mag_def': ('🛡️ Маг.защита', 'маг_защита'),
    'attack_speed': ('🏹 Скор.атаки', 'скорость_атаки'),
    'move_speed': ('🏃 Скор.движения', 'скорость_передвижения')
}

def check_db_structure():
    """Функция для проверки структуры базы данных"""
    try:
        conn = sqlite3.connect(DB_PATH)  # Используем константу DB_PATH
        cursor = conn.cursor()
        
        # Получаем список всех таблиц
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        
        logger.info("Найденные таблицы:")
        for table in tables:
            logger.info(f"Table: {table[0]}")
            cursor.execute(f"PRAGMA table_info('{table[0]}')")
            columns = cursor.fetchall()
            logger.info(f"Структура таблицы {table[0]}:")
            for column in columns:
                logger.info(f"  Column: {column[1]}, Type: {column[2]}")
            
        conn.close()
        return tables
        
    except Exception as e:
        logger.error(f"Ошибка при проверке структуры БД: {e}")
        return None

# Запустите эту функцию перед реализацией основного функционала
columns = check_db_structure()

def get_heroes_by_stat(stat, level):
    """Получение отсортированного списка героев по характеристике"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        growing_stats = {
            'ОЗ': ('ОЗ', 'прирост_ОЗ'),
            'реген_ОЗ': ('реген_ОЗ', 'прирост_реген_ОЗ'),
            'мана/энергия': ('мана/энергия', 'прирост_мана/энергия'),
            'реген_маны/энергии': ('реген_маны/энергии', 'прирост_реген_маны/энергии'),
            'физ_атака': ('физ_атака', 'прирост_физ_атака'),
            'физ_защита': ('физ_защита', 'прирост_физ_защита'),
            'маг_защита': ('маг_защита', 'прирост_маг_защита'),
            'скорость_атаки': ('скорость_атаки', 'прирост_скорость_атаки')
        }

        if stat in growing_stats:
            base_stat, growth_stat = growing_stats[stat]
            query = f"""
            SELECT n.имя,
                   no.{base_stat} + (up.{growth_stat} * (? - 1)) as calculated_stat
            FROM hero_names n
            JOIN hero_chars_no_up no ON n.id = no.character_id
            JOIN hero_chars_up up ON n.id = up.character_id
            ORDER BY calculated_stat DESC
            """
        else:
            query = f"""
            SELECT n.имя, no.{stat}
            FROM hero_names n
            JOIN hero_chars_no_up no ON n.id = no.character_id
            ORDER BY no.{stat} DESC
            """

        cursor.execute(query, (level,) if stat in growing_stats else ())
        results = cursor.fetchall()
        conn.close()
        return results

    except Exception as e:
        logger.error(f"Ошибка при получении данных из БД: {e}")
        return []

def register_handlers(bot):
    @bot.message_handler(commands=['chars_table'])
    def chars_table_cmd(message):
        try:
            kb = types.InlineKeyboardMarkup(row_width=2)
            for short_code, (btn_text, full_stat_name) in CHARACTERISTICS_MAP.items():
                kb.add(types.InlineKeyboardButton(
                    btn_text, 
                    callback_data=f"charTable::{short_code}"
                ))
            bot.send_message(
                message.chat.id,
                "Выберите характеристику для сравнения героев:",
                reply_markup=kb
            )
        except Exception as e:
            logger.error(f"Ошибка при отправке меню характеристик: {e}")
            bot.reply_to(message, "Произошла ошибка. Попробуйте позже.")

    @bot.callback_query_handler(func=lambda call: call.data.startswith("charTable::"))
    def char_table_callback(call):
        try:
            short_code = call.data.split("::")[1]
            stat_name = CHARACTERISTICS_MAP[short_code][1]
            
            ask_output_format(bot, call.message, stat_name)
            
        except Exception as e:
            logger.error(f"Ошибка в обработке callback: {e}")
            bot.answer_callback_query(
                call.id,
                "Произошла ошибка. Попробуйте позже."
            )

    @bot.callback_query_handler(func=lambda call: call.data.startswith('page:'))
    def handle_pagination(call):
        try:
            _, stat_name, level, page = call.data.split(':')
            page = int(page)
            level = int(level)
            
            results = get_heroes_by_stat(stat_name, level)
            if results:
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=create_formatted_message(results, page, stat_name, level),
                    reply_markup=create_pagination_keyboard(
                        page,
                        (len(results) + 29) // 30,
                        stat_name,
                        level
                    ),
                    parse_mode='HTML'  # Меняем на HTML
                )
                    
            bot.answer_callback_query(call.id)
                
        except Exception as e:
            logger.error(f"Ошибка при обработке пагинации: {e}")
            bot.answer_callback_query(
                call.id,
                "Произошла ошибка. Попробуйте позже."
            )

    @bot.callback_query_handler(func=lambda call: call.data.startswith("format:"))
    def format_handler(call):
        try:
            format_type, stat_name = call.data.split(":")[1:]
            msg = bot.send_message(
                call.message.chat.id,
                "Введите уровень героя (1-15):"
            )
            bot.register_next_step_handler(
                msg, 
                lambda m: process_level_input_with_format(m, bot, stat_name, format_type)
            )
        except Exception as e:
            logger.error(f"Ошибка при обработке формата: {e}")
            bot.answer_callback_query(call.id, "Произошла ошибка. Попробуйте позже.")

    @bot.callback_query_handler(func=lambda call: call.data.startswith("download_full:"))
    def handle_full_list_download(call):
        try:
            _, stat_name, level = call.data.split(":")
            level = int(level)
            results = get_heroes_by_stat(stat_name, level)
            
            if results:
                # Создаем текстовый файл с полным списком
                output = io.StringIO()
                output.write(f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне:\n\n")
                
                max_name_length = max(len(format_hero_name(hero)) for hero, _ in results)
                # Записываем заголовок
                output.write(f"{'№':3} │ {'Герой':<{max_name_length}} │ {'Значение':>10}\n")
                output.write("─" * (max_name_length + 17) + "\n")
                
                # Записываем все данные
                for idx, (hero, value) in enumerate(results, 1):
                    hero_name = format_hero_name(hero)
                    if value is not None:
                        value_str = f"{round(float(value), 2):>10.2f}"
                    else:
                        value_str = "нет данных"
                    output.write(f"{idx:3} │ {hero_name:<{max_name_length}} │ {value_str:>10}\n")
                
                output.seek(0)
                
                # Отправляем файл
                bot.send_document(
                    call.message.chat.id,
                    ('full_list.txt', output.getvalue().encode('utf-8')),
                    caption=f"📄 Полный список героев по '{stat_name}' на {level} уровне"
                )
                
            bot.answer_callback_query(call.id)
            
        except Exception as e:
            logger.error(f"Ошибка при скачивании полного списка: {e}")
            bot.answer_callback_query(
                call.id,
                "Произошла ошибка. Попробуйте позже."
            )

    def handle_text_pagination(bot, call):
        try:
            _, stat_name, level, page = call.data.split(':')
            level = int(level)
            page = int(page)
            
            results = get_heroes_by_stat(stat_name, level)
            if results:
                # Редактируем существующее сообщение вместо отправки нового
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=create_text_page(results, stat_name, level, page),
                    parse_mode='HTML',
                    reply_markup=create_text_pagination_markup(
                        len(results),
                        page,
                        stat_name,
                        level
                    )
                )
            
            bot.answer_callback_query(call.id)
            
        except Exception as e:
            logger.error(f"Ошибка при обработке пагинации: {e}")
            bot.answer_callback_query(
                call.id,
                "Произошла ошибка. Попробуйте позже."
            )

    @bot.callback_query_handler(func=lambda call: call.data.startswith('text_page:'))
    def text_page_handler(call):
        handle_text_pagination(bot, call)

    return chars_table_cmd

def create_pagination_keyboard(current_page, total_pages, stat, level):
    kb = types.InlineKeyboardMarkup(row_width=3)
    buttons = []
    
    if current_page > 1:
        buttons.append(types.InlineKeyboardButton(
            '⬅️', callback_data=f'page:{stat}:{level}:{current_page-1}'
        ))
    
    buttons.append(types.InlineKeyboardButton(
        f'{current_page}/{total_pages}', callback_data='current_page'
    ))
    
    if current_page < total_pages:
        buttons.append(types.InlineKeyboardButton(
            '➡️', callback_data=f'page:{stat}:{level}:{current_page+1}'
        ))
    
    kb.add(*buttons)
    return kb

def format_hero_name(name):
    """Форматирование имени героя с большой буквы"""
    return name.title()

def show_heroes_page(bot, chat_id, results, page, stat_name, level):
    HEROES_PER_PAGE = 30  # Увеличиваем количество героев на странице
    total_pages = (len(results) + HEROES_PER_PAGE - 1) // HEROES_PER_PAGE
    
    start_idx = (page - 1) * HEROES_PER_PAGE
    end_idx = start_idx + HEROES_PER_PAGE
    page_results = results[start_idx:end_idx]
    
    # Находим максимальную длину имени для выравнивания
    max_name_length = max(len(format_hero_name(hero)) for hero, _ in page_results)
    
    response = f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне:\n\n"
    response += "```\n"  # Beginning of monospace text block
    
    # Добавляем заголовок таблицы
    response += f"{'№':3} | {'Герой':<{max_name_length}} | {'Значение':>10}\n"
    response += "-" * (max_name_length + 17) + "\n"
    
    for idx, (hero, value) in enumerate(page_results, start=start_idx + 1):
        if value is not None:
            # Форматируем строку с выравниванием
            hero_name = format_hero_name(hero)
            value_str = f"{round(float(value), 2):>10.2f}"
            response += f"{idx:3} | {hero_name:<{max_name_length}} | {value_str}\n"
        else:
            response += f"{idx:3} | {hero_name:<{max_name_length}} | {'Нет данных':>10}\n"
    
    response += "```"  # End of monospace text block
    
    kb = create_pagination_keyboard(page, total_pages, stat_name, level)
    bot.send_message(
        chat_id,
        response,
        reply_markup=kb,
        parse_mode='Markdown'  # Включаем поддержку Markdown
    )

def create_formatted_message(results, stat_name, level):
    """Создание отформатированного сообщения со всеми героями в одной таблице"""
    # Находим максимальную длину имени
    max_name_length = max(len(format_hero_name(hero)) for hero, _ in results)
    
    message = f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне:\n\n"
    message += "<pre>"
    
    # Заголовок таблицы
    message += f"{'№':3} │ {'Герой':<{max_name_length}} │ {'Значение':>10}\n"
    message += "─" * 3 + "┼" + "─" * (max_name_length + 2) + "┼" + "─" * 12 + "\n"
    
    # Выводим всех героев
    for idx, (hero, value) in enumerate(results, start=1):
        hero_name = format_hero_name(hero)
        if value is not None:
            value_str = f"{round(float(value), 2):>10.2f}"
        else:
            value_str = "нет данных"
            
        message += f"{idx:3} │ {hero_name:<{max_name_length}} │ {value_str:>10}\n"
    
    message += "</pre>"
    return message

def create_table_image(results, stat_name, level):
    # Константы для размеров и отступов
    PADDING = 50
    LINE_HEIGHT = 25
    FONT_SIZE = 18
    COLUMNS = 3
    COLUMN_SPACING = 80
    
    # Константы для ширины столбцов внутри колонки
    NUM_COL_WIDTH = 40
    HERO_COL_WIDTH = max(len(format_hero_name(hero)) for hero, _ in results) * 11
    VAL_COL_WIDTH = 80
    COLUMN_WIDTH = NUM_COL_WIDTH + HERO_COL_WIDTH + VAL_COL_WIDTH + 20
    
    # Цветовая схема в стиле MLBB
    COLORS = {
        'background': '#1A1A2E',     # Темно-синий фон
        'header_bg': '#E94560',      # Красный для заголовков
        'header_text': '#FFFFFF',     # Белый для текста заголовков
        'text': '#FFFFFF',           # Белый для основного текста
        'value': '#FED7D7',          # Светло-красный для значений
        'column_bg': (22, 33, 62, 230),  # Темно-синий фон колонок с прозрачностью
        'row_alt': (31, 48, 94, 230),    # Альтернативный цвет строк с прозрачностью
        'border': '#E94560',         # Красный для границ
    }

    # Подготовка изображения
    heroes_per_column = (len(results) + COLUMNS - 1) // COLUMNS
    column_data = [results[i:i + heroes_per_column] for i in range(0, len(results), heroes_per_column)]
    
    width = (COLUMN_WIDTH * COLUMNS) + (COLUMN_SPACING * (COLUMNS - 1)) + (PADDING * 2)
    height = (heroes_per_column + 3) * LINE_HEIGHT + PADDING * 2

    # Загрузка случайного фона
    background_path = os.path.join(os.path.dirname(__file__), '..', 'media', 'background')
    background_files = [f for f in os.listdir(background_path) if f.endswith(('.png', '.jpg', '.jpeg'))]
    
    if (background_files):
        random_bg = random.choice(background_files)
        background = Image.open(os.path.join(background_path, random_bg))
        # Изменяем размер фона под размер нашего изображения
        background = background.resize((width, height), Image.Resampling.LANCZOS)
        # Создаем новое изображение с альфа-каналом
        img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
    else:
        # Если фоны не найдены, создаем сплошной фон
        img = Image.new('RGBA', (width, height), COLORS['background'])
        background = None

    draw = ImageDraw.Draw(img)
    
    try:
        # Пробуем загрузить Arial
        font_path = os.path.join(os.path.dirname(__file__), '..', 'assets', 'fonts', 'arial.ttf')
        title_font_path = os.path.join(os.path.dirname(__file__), '..', 'assets', 'fonts', 'arialbd.ttf')
        
        font = ImageFont.truetype(font_path, FONT_SIZE)
        title_font = ImageFont.truetype(title_font_path, FONT_SIZE + 4)
        
    except Exception as e:
        logger.error(f"Ошибка загрузки шрифта: {e}")
        # Если Arial не удалось загрузить, пробуем DejaVu
        try:
            font_path = os.path.join(os.path.dirname(__file__), '..', 'assets', 'fonts', 'DejaVuSansCondensed.ttf')
            font = ImageFont.truetype(font_path, FONT_SIZE)
            title_font = ImageFont.truetype(font_path, FONT_SIZE + 4)
        except:
            # Если и это не удалось, используем системный шрифт
            font = ImageFont.load_default()
            title_font = font

    # Рисуем заголовок с полупрозрачным фоном
    title = f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне"
    title_width = draw.textlength(title, font=title_font)
    
    # Фон для заголовка
    title_bg_height = LINE_HEIGHT + PADDING
    draw.rectangle(
        [(PADDING//2, PADDING//2),
         (width - PADDING//2, PADDING + title_bg_height)],
        fill=(22, 33, 62, 230)
    )
    
    draw.text(
        ((width - title_width) // 2, PADDING),
        title,
        font=title_font,
        fill=COLORS['header_text']
    )

    # Рисуем колонки
    y_start = PADDING + LINE_HEIGHT * 2
    for col_idx, column in enumerate(column_data):
        x_start = PADDING + col_idx * (COLUMN_WIDTH + COLUMN_SPACING)
        y = y_start
        
        x_num = x_start + 10
        x_hero = x_num + NUM_COL_WIDTH
        x_val = x_hero + HERO_COL_WIDTH
        
        # Фон колонки с прозрачностью
        draw.rectangle(
            [(x_start - PADDING//4, y - PADDING//4),
             (x_start + COLUMN_WIDTH - PADDING//4, height - PADDING)],
            fill=COLORS['column_bg'],
            outline=COLORS['border'],
            width=2
        )
        
        # Заголовок колонки
        draw.rectangle(
            [(x_start - PADDING//4, y - PADDING//4),
             (x_start + COLUMN_WIDTH - PADDING//4, y + LINE_HEIGHT)],
            fill=COLORS['header_bg']
        )
        
        draw.text((x_num, y), "№", font=font, fill=COLORS['header_text'])
        draw.text((x_hero, y), "Герой", font=font, fill=COLORS['header_text'])
        draw.text((x_val, y), "Значение", font=font, fill=COLORS['header_text'])
        
        y += LINE_HEIGHT

        start_idx = col_idx * heroes_per_column
        for idx, (hero, value) in enumerate(column, start=start_idx + 1):
            if idx % 2 == 0:
                draw.rectangle(
                    [(x_start - PADDING//4, y),
                     (x_start + COLUMN_WIDTH - PADDING//4, y + LINE_HEIGHT)],
                    fill=COLORS['row_alt']
                )
            
            hero_name = format_hero_name(hero)
            value_str = f"{round(float(value), 2):.2f}" if value is not None else "н/д"
            
            draw.text((x_num, y), str(idx).rjust(3), font=font, fill=COLORS['text'])
            draw.text((x_hero, y), hero_name, font=font, fill=COLORS['text'])
            draw.text((x_val, y), value_str.rjust(8), font=font, fill=COLORS['value'])
            
            y += LINE_HEIGHT

    # Объединяем фон и таблицу
    if background:
        final_image = Image.alpha_composite(background.convert('RGBA'), img)
    else:
        final_image = img

    # Сохранение
    img_byte_arr = io.BytesIO()
    final_image.save(img_byte_arr, format='PNG', quality=95)
    img_byte_arr.seek(0)
    return img_byte_arr

def process_level_input(message, bot, stat_name):
    try:
        level = int(message.text.strip())
        if not 1 <= level <= 15:
            raise ValueError("Уровень должен быть от 1 до 15")
            
        wait_msg = bot.send_message(
            message.chat.id,
            "⌛ Загрузка данных..."
        )
        
        results = get_heroes_by_stat(stat_name, level)
        
        if not results:
            bot.delete_message(message.chat.id, wait_msg.message_id)
            bot.reply_to(message, "Не удалось получить данные героев.")
            return
            
        # Создаем изображение
        img_bytes = create_table_image(results, stat_name, level)
        
        bot.delete_message(message.chat.id, wait_msg.message_id)
        bot.send_photo(
            message.chat.id,
            img_bytes,
            caption=f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне"
        )
        
    except ValueError:
        msg = bot.reply_to(
            message,
            "Пожалуйста, введите корректный уровень (от 1 до 15)."
        )
        bot.register_next_step_handler(msg, lambda m: process_level_input(m, bot, stat_name))
    except Exception as e:
        logger.error(f"Ошибка при обработке уровня: {e}")
        bot.reply_to(message, "Произошла ошибка. Попробуйте позже.")

def process_level_input_with_cancel(message, bot, stat_name):
    if message.text == '❌ Отмена':
        bot.send_message(
            message.chat.id, 
            "Операция отменена",
            reply_markup=types.ReplyKeyboardRemove()
        )
        return
    
    process_level_input(message, bot, stat_name)

def ask_output_format(bot, message, stat_name):
    """Спрашиваем у пользователя желаемый формат вывода"""
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("📝 Текст", callback_data=f"format:text:{stat_name}"),
        types.InlineKeyboardButton("🖼 Изображение", callback_data=f"format:image:{stat_name}"),
        types.InlineKeyboardButton("📊 Excel", callback_data=f"format:excel:{stat_name}"),
        types.InlineKeyboardButton("📄 PDF", callback_data=f"format:pdf:{stat_name}")
    )
    bot.send_message(
        message.chat.id,
        "Выберите формат вывода данных:",
        reply_markup=markup
    )

def create_excel_file(results, stat_name, level):
    """Создание Excel файла"""
    df = pd.DataFrame(results, columns=['Герой', 'Значение'])
    df.index = df.index + 1  # Начинаем нумерацию с 1
    
    # Создаем файл Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'Рейтинг_{stat_name}')
        worksheet = writer.sheets[f'Рейтинг_{stat_name}']
        
        # Форматирование
        for idx, col in enumerate(df.columns):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 2)].width = 20
    
    output.seek(0)
    return output

def create_pdf_file(results, stat_name, level):
    """Создание PDF файла с табличными данными - все герои на одной странице"""
    try:
        # Используем fpdf2 вместо fpdf для лучшей поддержки Unicode
        from fpdf import FPDF
        import os
        import io
        import math

        # Определяем путь к шрифтам в папке assets
        font_dir = os.path.join(os.path.dirname(__file__), '..', 'assets', 'fonts')
        
        # Проверяем наличие шрифтов
        font_files = os.listdir(font_dir) if os.path.exists(font_dir) else []
        logger.info(f"Найденные шрифты: {font_files}")
        
        # Создаем PDF с поддержкой Unicode
        class PDF(FPDF):
            def __init__(self):
                super().__init__(orientation='P', unit='mm', format='A4')
                # Устанавливаем автоматический разрыв страниц
                self.set_auto_page_break(auto=True, margin=5)
                
            def header(self):
                # Заголовок с логотипом MLBB
                self.set_font('DejaVu', 'B', 12)
                self.set_text_color(220, 50, 50)  # Красный цвет для заголовка
                title = f"Рейтинг героев по '{stat_name}' на {level} уровне"
                self.cell(0, 6, txt=title, ln=True, align='C')
                
            def footer(self):
                # Нижний колонтитул
                self.set_y(-7)
                self.set_font('DejaVu', 'I', 6)
                self.set_text_color(128, 128, 128)
                self.cell(0, 5, 'MLBB Helper Bot', 0, 0, 'C')
                
        # Создаем PDF-документ
        pdf = PDF()
        
        # Добавляем шрифт DejaVu для поддержки кириллицы
        dejavu_path = os.path.join(font_dir, 'DejaVuSansCondensed.ttf')
        pdf.add_font('DejaVu', '', dejavu_path, uni=True)
        pdf.add_font('DejaVu', 'B', os.path.join(font_dir, 'DejaVuSansCondensed-Bold.ttf'), uni=True)
        pdf.add_font('DejaVu', 'I', os.path.join(font_dir, 'DejaVuSansCondensed-Oblique.ttf'), uni=True)
        
        pdf.add_page()
        
        # Используем шрифт DejaVu с поддержкой Unicode, но меньшего размера
        pdf.set_font('DejaVu', '', 6)  # Уменьшаем размер шрифта для компактности
        
        # Определяем количество колонок и строк для размещения всех героев
        total_heroes = len(results)
        columns = 3  # Используем 3 колонки для компактности
        
        # Вычисляем количество строк в каждой колонке
        rows_per_column = math.ceil(total_heroes / columns)
        
        # Определяем ширину колонок (A4 = 210mm ширина)
        page_width = 210
        margin = 5  # Уменьшаем отступы для большей компактности
        usable_width = page_width - 2 * margin
        column_width = usable_width / columns
        
        # Определяем ширину столбцов внутри колонки
        num_width = 7
        value_width = 18
        hero_width = column_width - num_width - value_width
        
        # Настройка цветов для таблицы
        header_color = (220, 50, 50)  # Красный для заголовка
        row1_color = (240, 240, 240)  # Светло-серый для четных строк
        row2_color = (255, 255, 255)  # Белый для нечетных строк
        text_color = (0, 0, 0)  # Черный для текста
        
        # Высота строки
        row_height = 4  # Уменьшаем высоту строки для компактности
        
        # Отступ сверху после заголовка
        y_offset = 10
        
        # Рисуем таблицы по колонкам
        for col in range(columns):
            # Вычисляем x-координату для текущей колонки
            x = margin + col * column_width
            
            # Заголовки таблицы
            pdf.set_xy(x, y_offset)
            pdf.set_fill_color(*header_color)
            pdf.set_text_color(255, 255, 255)  # Белый текст для заголовка
            pdf.cell(num_width, row_height, '№', 1, 0, 'C', True)
            pdf.cell(hero_width, row_height, 'Герой', 1, 0, 'C', True)
            pdf.cell(value_width, row_height, 'Значение', 1, 1, 'C', True)
            
            # Данные таблицы
            pdf.set_text_color(*text_color)
            
            # Определяем диапазон индексов для текущей колонки
            start_idx = col * rows_per_column
            end_idx = min(start_idx + rows_per_column, total_heroes)
            
            for i in range(start_idx, end_idx):
                idx = i + 1
                hero, value = results[i]
                
                # Чередуем цвета строк
                if idx % 2 == 0:
                    pdf.set_fill_color(*row1_color)
                else:
                    pdf.set_fill_color(*row2_color)
                
                hero_name = format_hero_name(hero)
                value_str = f"{round(float(value), 2):.2f}" if value is not None else "н/д"
                
                # Устанавливаем позицию для текущей строки
                current_y = y_offset + row_height + (i - start_idx) * row_height
                pdf.set_xy(x, current_y)
                
                # Рисуем ячейки
                pdf.cell(num_width, row_height, str(idx), 1, 0, 'C', True)
                pdf.cell(hero_width, row_height, hero_name, 1, 0, 'L', True)
                pdf.cell(value_width, row_height, value_str, 1, 1, 'R', True)
        
        # Сохраняем PDF в байтовый поток
        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
        
    except Exception as e:
        logger.error(f"Ошибка при создании PDF: {e}")
        raise

def process_level_input_with_format(message, bot, stat_name, format_type):
    """Обработка ввода уровня с учетом выбранного формата"""
    try:
        level = int(message.text.strip())
        if not 1 <= level <= 15:
            msg = bot.reply_to(
                message,
                "Пожалуйста, введите корректный уровень (от 1 до 15)."
            )
            bot.register_next_step_handler(
                msg, 
                lambda m: process_level_input_with_format(m, bot, stat_name, format_type)
            )
            return
            
        wait_msg = bot.send_message(
            message.chat.id,
            "⌛ Загрузка данных..."
        )
        
        results = get_heroes_by_stat(stat_name, level)
        if not results:
            bot.delete_message(message.chat.id, wait_msg.message_id)
            bot.reply_to(message, "Не удалось получить данные героев.")
            return

        bot.delete_message(message.chat.id, wait_msg.message_id)
        
        format_handlers = {
            "text": send_text_format,
            "image": send_image_format,
            "excel": send_excel_format,
            "pdf": send_pdf_format
        }
        
        if format_type in format_handlers:
            format_handlers[format_type](bot, message, results, stat_name, level)
        else:
            bot.reply_to(message, "Неизвестный формат. Попробуйте снова.")
            
    except ValueError:
        msg = bot.reply_to(
            message,
            "Пожалуйста, введите корректный уровень (от 1 до 15)."
        )
        bot.register_next_step_handler(
            msg, 
            lambda m: process_level_input_with_format(m, bot, stat_name, format_type)
        )
    except Exception as e:
        logger.error(f"Ошибка при обработке уровня: {e}")
        bot.reply_to(message, "Произошла ошибка. Попробуйте позже.")

# Функции отправки в разных форматах
def send_text_format(bot, message, results, stat_name, level, page=1):
    """Отправка данных в текстовом формате с пагинацией"""
    ITEMS_PER_PAGE = 30
    total_pages = (len(results) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    
    # Получаем данные для текущей страницы
    start_idx = (page - 1) * ITEMS_PER_PAGE
    end_idx = start_idx + ITEMS_PER_PAGE
    page_results = results[start_idx:end_idx]
    
    # Находим максимальную длину имени для выравнивания
    max_name_length = max(len(format_hero_name(hero)) for hero, _ in results)
    
    response = f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне (стр. {page}/{total_pages}):\n\n"
    response += "<pre>"
    
    # Добавляем заголовок таблицы
    response += f"{'№':3} │ {'Герой':<{max_name_length}} │ {'Значение':>10}\n"
    response += "─" * (max_name_length + 17) + "\n"
    
    # Выводим данные текущей страницы
    for idx, (hero, value) in enumerate(page_results, start=start_idx + 1):
        hero_name = format_hero_name(hero)
        if value is not None:
            value_str = f"{round(float(value), 2):>10.2f}"
        else:
            value_str = "нет данных"
            
        response += f"{idx:3} │ {hero_name:<{max_name_length}} │ {value_str:>10}\n"
    
    response += "</pre>"
    
    # Создаем кнопки пагинации
    markup = types.InlineKeyboardMarkup()
    nav_buttons = []
    
    if page > 1:
        nav_buttons.append(types.InlineKeyboardButton(
            '◀️', callback_data=f'text_page:{stat_name}:{level}:{page-1}'
        ))
    
    nav_buttons.append(types.InlineKeyboardButton(
        f'{page}/{total_pages}', callback_data='current_page'
    ))
    
    if page < total_pages:
        nav_buttons.append(types.InlineKeyboardButton(
            '▶️', callback_data=f'text_page:{stat_name}:{level}:{page+1}'
        ))
    
    markup.add(*nav_buttons)
    
    bot.send_message(
        message.chat.id,
        response,
        parse_mode='HTML',
        reply_markup=markup
    )

def handle_text_pagination(bot, call):
    try:
        _, stat_name, level, page = call.data.split(':')
        level = int(level)
        page = int(page)
        
        results = get_heroes_by_stat(stat_name, level)
        if results:
            # Редактируем существующее сообщение вместо отправки нового
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=create_text_page(results, stat_name, level, page),
                parse_mode='HTML',
                reply_markup=create_text_pagination_markup(
                    len(results),
                    page,
                    stat_name,
                    level
                )
            )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        logger.error(f"Ошибка при обработке пагинации: {e}")
        bot.answer_callback_query(
            call.id,
            "Произошла ошибка. Попробуйте позже."
        )
        bot.answer_callback_query(
            call.id,
            "Произошла ошибка. Попробуйте позже."
        )

def create_text_page(results, stat_name, level, page):
    """Создание текстовой страницы с данными"""
    ITEMS_PER_PAGE = 30
    total_pages = (len(results) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    
    start_idx = (page - 1) * ITEMS_PER_PAGE
    end_idx = start_idx + ITEMS_PER_PAGE
    page_results = results[start_idx:end_idx]
    
    max_name_length = max(len(format_hero_name(hero)) for hero, _ in results)
    
    response = f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне (стр. {page}/{total_pages}):\n\n"
    response += "<pre>"
    # Заголовок таблицы
    response += f"{'№':3} │ {'Герой':<{max_name_length}} │ {'Значение':>10}\n"
    # Сплошная разделительная линия
    response += "─" * (max_name_length + 17) + "\n"
    
    for idx, (hero, value) in enumerate(page_results, start=start_idx + 1):
        hero_name = format_hero_name(hero)
        if value is not None:
            value_str = f"{round(float(value), 2):>10.2f}"
        else:
            value_str = "нет данных"
            
        response += f"{idx:3} │ {hero_name:<{max_name_length}} │ {value_str:>10}\n"
    
    response += "</pre>"
    return response

def create_text_pagination_markup(total_items, current_page, stat_name, level):
    """Создание клавиатуры с кнопками пагинации"""
    markup = types.InlineKeyboardMarkup()
    nav_buttons = []
    total_pages = (total_items + 29) // 30
    
    if current_page > 1:
        nav_buttons.append(types.InlineKeyboardButton(
            '◀️', callback_data=f'text_page:{stat_name}:{level}:{current_page-1}'
        ))
    
    nav_buttons.append(types.InlineKeyboardButton(
        f'{current_page}/{total_pages}', callback_data='current_page'
    ))
    
    if current_page < total_pages:
        nav_buttons.append(types.InlineKeyboardButton(
            '▶️', callback_data=f'text_page:{stat_name}:{level}:{current_page+1}'
        ))
    
    markup.add(*nav_buttons)
    return markup

def send_image_format(bot, message, results, stat_name, level):
    img_bytes = create_table_image(results, stat_name, level)
    bot.send_document(
        message.chat.id,
        ('table.png', img_bytes),
        caption=f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне"
    )

def send_excel_format(bot, message, results, stat_name, level):
    excel_bytes = create_excel_file(results, stat_name, level)
    bot.send_document(
        message.chat.id,
        ('rating.xlsx', excel_bytes),
        caption=f"📊 Рейтинг героев по '{stat_name}' на {level} уровне"
    )

def send_pdf_format(bot, message, results, stat_name, level):
    """Отправка PDF файла с данными"""
    try:
        chat_id = message.chat.id
        pdf_data = create_pdf_file(results, stat_name, level)
        
        # Создаем временное имя файла для отправки
        filename = f"heroes_rating_{stat_name}_{level}.pdf"
        
        # Отправляем PDF как документ
        bot.send_document(
            chat_id,
            pdf_data,
            caption=f"🏆 Рейтинг героев по '{stat_name}' на {level} уровне",
            visible_file_name=filename
        )
        
        logger.info(f"PDF файл успешно отправлен пользователю {message.from_user.id}")
    except Exception as e:
        logger.error(f"Ошибка при отправке PDF: {e}")
        bot.reply_to(
            message,
            "Произошла ошибка при создании PDF. Попробуйте другой формат.",
            reply_parameters=telebot.types.ReplyParameters(message_id=message.message_id)
        )
